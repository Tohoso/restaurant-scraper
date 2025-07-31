#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
飲食店データ統合・Excel出力機能
食べログとホットペッパーのデータを統合してExcelファイルに出力
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Optional
import logging
from datetime import datetime
import re

# ログ設定
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class RestaurantDataIntegrator:
    """飲食店データ統合クラス"""
    
    def __init__(self):
        """初期化"""
        self.restaurants = []
        
    def add_restaurants(self, restaurant_list: List[Dict]):
        """
        飲食店データを追加
        
        Args:
            restaurant_list (List[Dict]): 飲食店データリスト
        """
        self.restaurants.extend(restaurant_list)
        logger.info(f"{len(restaurant_list)}件の飲食店データを追加しました")
    
    def clean_phone_number(self, phone: str) -> str:
        """
        電話番号を正規化
        
        Args:
            phone (str): 電話番号
            
        Returns:
            str: 正規化された電話番号
        """
        if not phone:
            return ""
        
        # 数字とハイフンのみ抽出
        phone_clean = re.sub(r'[^\d\-]', '', phone)
        
        # 先頭の0を確保
        if phone_clean and not phone_clean.startswith('0'):
            # 03、06、011などの市外局番パターンを検出
            if len(phone_clean) >= 10:
                phone_clean = '0' + phone_clean
        
        return phone_clean
    
    def clean_address(self, address: str) -> str:
        """
        住所を正規化
        
        Args:
            address (str): 住所
            
        Returns:
            str: 正規化された住所
        """
        if not address:
            return ""
        
        # 不要な文字列を除去
        address_clean = address.strip()
        
        # 移転情報などを除去
        if 'このお店は' in address_clean:
            address_clean = address_clean.split('このお店は')[0].strip()
        
        if '※' in address_clean:
            address_clean = address_clean.split('※')[0].strip()
        
        return address_clean
    
    def remove_duplicates(self):
        """重複データを除去"""
        unique_restaurants = []
        seen_names = set()
        seen_phones = set()
        
        for restaurant in self.restaurants:
            name = restaurant.get('shop_name', '').strip()
            phone = self.clean_phone_number(restaurant.get('phone', ''))
            
            # 店名または電話番号で重複チェック
            is_duplicate = False
            
            if name and name in seen_names:
                is_duplicate = True
            
            if phone and phone in seen_phones:
                is_duplicate = True
            
            if not is_duplicate:
                if name:
                    seen_names.add(name)
                if phone:
                    seen_phones.add(phone)
                unique_restaurants.append(restaurant)
        
        removed_count = len(self.restaurants) - len(unique_restaurants)
        self.restaurants = unique_restaurants
        
        logger.info(f"重複除去: {removed_count}件を除去、{len(unique_restaurants)}件が残りました")
    
    def validate_data(self):
        """データの品質チェック"""
        valid_restaurants = []
        
        for restaurant in self.restaurants:
            name = restaurant.get('shop_name', '').strip()
            phone = self.clean_phone_number(restaurant.get('phone', ''))
            address = self.clean_address(restaurant.get('address', ''))
            
            # 必須項目チェック
            if name:  # 店名は必須
                # 電話番号と住所を正規化して更新
                restaurant['phone'] = phone
                restaurant['address'] = address
                valid_restaurants.append(restaurant)
            else:
                logger.warning(f"店名が空のデータをスキップ: {restaurant}")
        
        invalid_count = len(self.restaurants) - len(valid_restaurants)
        self.restaurants = valid_restaurants
        
        logger.info(f"データ検証: {invalid_count}件の無効データを除去、{len(valid_restaurants)}件が有効")
    
    def create_excel_report(self, filename: str = None) -> str:
        """
        Excelレポートを作成
        
        Args:
            filename (str): ファイル名（省略時は自動生成）
            
        Returns:
            str: 作成されたファイルのパス
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"飲食店営業リスト_{timestamp}.xlsx"
        
        # データフレーム作成
        df_data = []
        for i, restaurant in enumerate(self.restaurants, 1):
            df_data.append({
                'No': i,
                '店名': restaurant.get('shop_name', ''),
                '電話番号': restaurant.get('phone', ''),
                '住所': restaurant.get('address', ''),
                'ジャンル': restaurant.get('genre', ''),
                '最寄り駅': restaurant.get('station', ''),
                '営業時間': restaurant.get('open_time', ''),
                'データソース': restaurant.get('source', ''),
                'URL': restaurant.get('url', ''),
                '調査日': datetime.now().strftime("%Y-%m-%d")
            })
        
        df = pd.DataFrame(df_data)
        
        # Excelファイル作成
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # メインデータシート
            df.to_excel(writer, sheet_name='飲食店リスト', index=False)
            
            # サマリーシート
            summary_data = {
                '項目': [
                    '総件数',
                    '電話番号あり',
                    '住所あり',
                    'ジャンルあり',
                    '食べログ',
                    'ホットペッパー',
                    '作成日時'
                ],
                '値': [
                    len(self.restaurants),
                    len([r for r in self.restaurants if r.get('phone')]),
                    len([r for r in self.restaurants if r.get('address')]),
                    len([r for r in self.restaurants if r.get('genre')]),
                    len([r for r in self.restaurants if r.get('source') == '食べログ']),
                    len([r for r in self.restaurants if r.get('source') == 'ホットペッパー']),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='サマリー', index=False)
        
        # スタイル適用
        self._apply_excel_styling(filename)
        
        logger.info(f"Excelファイルを作成しました: {filename}")
        return filename
    
    def _apply_excel_styling(self, filename: str):
        """Excelファイルにスタイルを適用"""
        wb = openpyxl.load_workbook(filename)
        
        # メインデータシートのスタイル
        ws_main = wb['飲食店リスト']
        
        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 罫線
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行にスタイル適用
        for cell in ws_main[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # データ行にスタイル適用
        for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # 列幅調整
        column_widths = {
            'A': 5,   # No
            'B': 30,  # 店名
            'C': 15,  # 電話番号
            'D': 40,  # 住所
            'E': 15,  # ジャンル
            'F': 20,  # 最寄り駅
            'G': 25,  # 営業時間
            'H': 12,  # データソース
            'I': 50,  # URL
            'J': 12   # 調査日
        }
        
        for col, width in column_widths.items():
            ws_main.column_dimensions[col].width = width
        
        # サマリーシートのスタイル
        ws_summary = wb['サマリー']
        
        # ヘッダー行にスタイル適用
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # データ行にスタイル適用
        for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
        
        # 列幅調整
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 25
        
        wb.save(filename)
    
    def get_statistics(self) -> Dict:
        """統計情報を取得"""
        stats = {
            'total_count': len(self.restaurants),
            'with_phone': len([r for r in self.restaurants if r.get('phone')]),
            'with_address': len([r for r in self.restaurants if r.get('address')]),
            'with_genre': len([r for r in self.restaurants if r.get('genre')]),
            'tabelog_count': len([r for r in self.restaurants if r.get('source') == '食べログ']),
            'hotpepper_count': len([r for r in self.restaurants if r.get('source') == 'ホットペッパー'])
        }
        
        return stats

def test_data_integrator():
    """データ統合機能のテスト"""
    print("🔧 データ統合機能 テスト開始")
    
    # テストデータ作成
    test_data = [
        {
            'shop_name': 'テスト居酒屋1',
            'phone': '03-1234-5678',
            'address': '東京都渋谷区渋谷1-1-1',
            'genre': '居酒屋',
            'station': '渋谷駅',
            'open_time': '17:00-24:00',
            'source': '食べログ',
            'url': 'https://example.com/1'
        },
        {
            'shop_name': 'テストレストラン2',
            'phone': '06-9876-5432',
            'address': '大阪府大阪市北区梅田1-1-1',
            'genre': 'イタリアン',
            'station': '梅田駅',
            'open_time': '11:00-22:00',
            'source': 'ホットペッパー',
            'url': 'https://example.com/2'
        }
    ]
    
    # データ統合テスト
    integrator = RestaurantDataIntegrator()
    integrator.add_restaurants(test_data)
    integrator.validate_data()
    integrator.remove_duplicates()
    
    # 統計情報表示
    stats = integrator.get_statistics()
    print(f"✅ 統計情報:")
    print(f"   総件数: {stats['total_count']}")
    print(f"   電話番号あり: {stats['with_phone']}")
    print(f"   住所あり: {stats['with_address']}")
    
    # Excelファイル作成
    filename = integrator.create_excel_report("test_restaurant_list.xlsx")
    print(f"✅ テスト用Excelファイルを作成: {filename}")

if __name__ == "__main__":
    test_data_integrator()

